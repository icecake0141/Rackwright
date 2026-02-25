"""
Copyright 2026 Rackwright Contributors
SPDX-License-Identifier: Apache-2.0

This file was created or modified with the assistance of an AI (Large Language Model).
Review required for correctness, security, and licensing.
"""

from __future__ import annotations

import os
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from rackwright.diff_service import diff
from rackwright.generation_service import generate
from rackwright.models import (
    ArtifactFile,
    Base,
    Cabling,
    DiffItem,
    Project,
    TemplateSection,
    TemplateSet,
)
from rackwright.template_services import create_project_from_template_set


def _new_session() -> Session:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    return Session(engine)


def _seed_project(session: Session) -> int:
    template_set = TemplateSet(name="diff-ts")
    session.add(template_set)
    session.flush()
    session.add(
        TemplateSection(
            template_set_id=template_set.id,
            target_type="Project",
            category="General",
            section_order=1,
            output_targets='["word", "excel"]',
            applicable_roles=None,
            text="Doc for {{ project.name }} {{ project.notes or '' }}",
        )
    )
    session.flush()

    project = create_project_from_template_set(
        session,
        project_name="diff-project",
        owner=None,
        notes=None,
        template_set_id=template_set.id,
    )
    session.flush()

    session.add(
        Cabling(
            project_id=project.id,
            a_device="srv1",
            a_port="eth0",
            a_port_type="dcim.interface",
            b_device="sw1",
            b_port="ge-0/0/1",
            b_port_type="dcim.interface",
            cable_type="cat6",
            label="L1",
            normalized_key="srv1::eth0::dcim.interface|sw1::ge-0/0/1::dcim.interface",
        )
    )
    session.commit()
    return project.id


def test_diff_detects_and_marks_known_excel_cell_change(tmp_path: Path) -> None:
    os.environ["RACKWRIGHT_DATA_DIR"] = str(tmp_path)
    with _new_session() as session:
        project_id = _seed_project(session)

        v1 = generate(session, project_id, "all", None, "v1")
        session.commit()

        cabling = session.query(Cabling).filter(Cabling.project_id == project_id).one()
        cabling.label = "L2"
        project = session.query(Project).filter(Project.id == project_id).one()
        project.notes = "word-changed"
        session.commit()

        v2 = generate(session, project_id, "all", None, "v2")
        session.commit()

        result = diff(session, v1.artifact_version.id, v2.artifact_version.id)
        session.commit()

        items = (
            session.query(DiffItem)
            .filter(DiffItem.diff_report_id == result.diff_report.id)
            .all()
        )
        assert len(items) >= 1
        assert any(item.artifact_type == "excel" for item in items)

        diff_excel_file = (
            session.query(ArtifactFile)
            .filter(
                ArtifactFile.artifact_version_id == v2.artifact_version.id,
                ArtifactFile.artifact_type == "diff_excel",
            )
            .one()
        )
        path = tmp_path / diff_excel_file.relative_path
        assert path.exists()

        wb = load_workbook(path)
        ws = wb["wiring"]
        target_cell = ws.cell(row=2, column=6)
        assert target_cell.value == "L2"
        assert target_cell.comment is not None

        diff_word_file = (
            session.query(ArtifactFile)
            .filter(
                ArtifactFile.artifact_version_id == v2.artifact_version.id,
                ArtifactFile.artifact_type == "diff_word",
            )
            .one()
        )
        word_path = tmp_path / diff_word_file.relative_path
        assert word_path.exists()

        doc = Document(word_path)
        assert len(list(doc.comments)) >= 1

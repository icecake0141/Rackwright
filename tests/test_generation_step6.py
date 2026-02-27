"""
Copyright 2026 Rackwright Contributors
SPDX-License-Identifier: Apache-2.0

This file was created or modified with the assistance of an AI (Large Language Model).
Review required for correctness, security, and licensing.
"""

from __future__ import annotations

import json
import os
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from rackwright.generation_service import generate
from rackwright.models import (
    ArtifactFile,
    ArtifactVersion,
    Base,
    Cabling,
    Device,
    Rack,
    SectionApplicationRule,
    SectionSnapshot,
    TemplateSection,
    TemplateSet,
)
from rackwright.template_services import create_project_from_template_set


def _new_session() -> Session:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    return Session(engine)


def _create_project_with_snapshot(session: Session) -> int:
    template_set = TemplateSet(name="gen-ts")
    session.add(template_set)
    session.flush()
    session.add(
        TemplateSection(
            template_set_id=template_set.id,
            target_type="Project",
            category="General",
            section_order=1,
            output_targets='["word", "excel"]',
            applicable_roles=None,
            text="Project {{ project.name }}",
        )
    )
    session.flush()

    project = create_project_from_template_set(
        session,
        project_name="gen-project",
        owner="owner",
        notes=None,
        template_set_id=template_set.id,
    )
    session.commit()
    return project.id


def test_generate_all_creates_files_and_db_rows(tmp_path: Path) -> None:
    os.environ["RACKWRIGHT_DATA_DIR"] = str(tmp_path)
    with _new_session() as session:
        project_id = _create_project_with_snapshot(session)
        result = generate(session, project_id, "all", None, "first")
        session.commit()

        assert result.skipped is False
        version = result.artifact_version
        assert version.version_number == 1
        assert version.success_word is True
        assert version.success_excel is True
        assert version.success_image is True

        files = (
            session.query(ArtifactFile)
            .filter(ArtifactFile.artifact_version_id == version.id)
            .all()
        )
        assert len(files) == 3
        paths = {Path(tmp_path / f.relative_path) for f in files}
        for path in paths:
            assert path.exists()
        assert (
            tmp_path / f"projects/{project_id}/versions/1/word/document.docx"
        ).exists()
        assert (
            tmp_path / f"projects/{project_id}/versions/1/excel/document.xlsx"
        ).exists()
        assert (
            tmp_path / f"projects/{project_id}/versions/1/images/rack_layout.svg"
        ).exists()


def test_generate_idempotency_returns_existing_version(tmp_path: Path) -> None:
    os.environ["RACKWRIGHT_DATA_DIR"] = str(tmp_path)
    with _new_session() as session:
        project_id = _create_project_with_snapshot(session)
        first = generate(session, project_id, "word", None, "run1")
        session.commit()

        second = generate(session, project_id, "word", None, "run1")
        session.commit()

        assert first.artifact_version.id == second.artifact_version.id
        assert second.skipped is True
        count = (
            session.query(ArtifactVersion)
            .filter(ArtifactVersion.project_id == project_id)
            .count()
        )
        assert count == 1


def test_partial_failure_still_creates_artifact_version(
    tmp_path: Path, monkeypatch
) -> None:
    os.environ["RACKWRIGHT_DATA_DIR"] = str(tmp_path)
    with _new_session() as session:
        project_id = _create_project_with_snapshot(session)

        import rackwright.generation_service as generation_service

        def fail_excel(*args, **kwargs):
            raise RuntimeError("excel fail")

        monkeypatch.setattr(generation_service, "_write_excel", fail_excel)

        result = generate(session, project_id, "all", None, "with-failure")
        session.commit()

        version = result.artifact_version
        assert version.success_word is True
        assert version.success_excel is False
        assert version.success_image is True
        assert version.errors_json is not None
        errors = json.loads(version.errors_json)
        assert any(e["type"] == "excel" for e in errors)


def test_section_filters_apply_role_and_rack_scope(tmp_path: Path) -> None:
    os.environ["RACKWRIGHT_DATA_DIR"] = str(tmp_path)
    with _new_session() as session:
        template_set = TemplateSet(name="filter-ts")
        session.add(template_set)
        session.flush()
        session.add(
            TemplateSection(
                template_set_id=template_set.id,
                target_type="Device",
                category="Devices",
                section_order=1,
                output_targets='["excel"]',
                applicable_roles=None,
                text="{{ target.name }}",
            )
        )
        session.flush()

        project = create_project_from_template_set(
            session,
            project_name="filter-project",
            owner=None,
            notes=None,
            template_set_id=template_set.id,
        )
        session.flush()

        rack_a = Rack(project_id=project.id, name="rack-a", rack_height_u=42)
        rack_b = Rack(project_id=project.id, name="rack-b", rack_height_u=42)
        session.add_all([rack_a, rack_b])
        session.flush()

        session.add_all(
            [
                Device(
                    project_id=project.id,
                    rack_id=rack_a.id,
                    name="srv-a",
                    role="Server",
                ),
                Device(
                    project_id=project.id,
                    rack_id=rack_b.id,
                    name="srv-b",
                    role="Server",
                ),
                Device(
                    project_id=project.id, rack_id=rack_a.id, name="sw-a", role="Switch"
                ),
            ]
        )
        session.flush()

        section_snapshot = session.query(SectionSnapshot).one()
        rule = (
            session.query(SectionApplicationRule)
            .filter(
                SectionApplicationRule.project_id == project.id,
                SectionApplicationRule.section_snapshot_id == section_snapshot.id,
            )
            .one()
        )
        rule.filters_json = json.dumps({"role": ["Server"], "rack_scope": ["rack-a"]})
        session.commit()

        result = generate(session, project.id, "excel", None, "filter-run")
        session.commit()

        excel_file = (
            session.query(ArtifactFile)
            .filter(
                ArtifactFile.artifact_version_id == result.artifact_version.id,
                ArtifactFile.artifact_type == "excel",
            )
            .one()
        )
        wb = load_workbook(tmp_path / excel_file.relative_path)
        checklist = wb["checklists"]
        values = [
            checklist.cell(row=i, column=2).value
            for i in range(2, checklist.max_row + 1)
        ]
        assert values == ["srv-a"]


def test_target_scope_switches_for_rack_and_device_with_project_context(
    tmp_path: Path,
) -> None:
    os.environ["RACKWRIGHT_DATA_DIR"] = str(tmp_path)
    with _new_session() as session:
        template_set = TemplateSet(name="scope-ts")
        session.add(template_set)
        session.flush()
        session.add_all(
            [
                TemplateSection(
                    template_set_id=template_set.id,
                    target_type="Rack",
                    category="Scope",
                    section_order=1,
                    output_targets='["excel"]',
                    applicable_roles=None,
                    text="R={{ target.name }} P={{ project.name }}",
                ),
                TemplateSection(
                    template_set_id=template_set.id,
                    target_type="Device",
                    category="Scope",
                    section_order=2,
                    output_targets='["excel"]',
                    applicable_roles=None,
                    text="D={{ target.name }} P={{ project.name }}",
                ),
            ]
        )
        session.flush()

        project = create_project_from_template_set(
            session,
            project_name="scope-project",
            owner=None,
            notes=None,
            template_set_id=template_set.id,
        )
        session.flush()

        rack_a = Rack(project_id=project.id, name="rack-a", rack_height_u=42)
        rack_b = Rack(project_id=project.id, name="rack-b", rack_height_u=42)
        session.add_all([rack_a, rack_b])
        session.flush()

        session.add_all(
            [
                Device(
                    project_id=project.id,
                    rack_id=rack_a.id,
                    name="srv-a",
                    role="Server",
                ),
                Device(
                    project_id=project.id,
                    rack_id=rack_b.id,
                    name="srv-b",
                    role="Server",
                ),
            ]
        )
        session.commit()

        result = generate(session, project.id, "excel", None, "scope-check")
        session.commit()

        excel_file = (
            session.query(ArtifactFile)
            .filter(
                ArtifactFile.artifact_version_id == result.artifact_version.id,
                ArtifactFile.artifact_type == "excel",
            )
            .one()
        )
        wb = load_workbook(tmp_path / excel_file.relative_path)
        checklist = wb["checklists"]
        values = [
            checklist.cell(row=i, column=2).value
            for i in range(2, checklist.max_row + 1)
        ]

        assert "R=rack-a P=scope-project" in values
        assert "R=rack-b P=scope-project" in values
        assert "D=srv-a P=scope-project" in values
        assert "D=srv-b P=scope-project" in values


def test_generate_excel_includes_field_operation_sheets(tmp_path: Path) -> None:
    os.environ["RACKWRIGHT_DATA_DIR"] = str(tmp_path)
    with _new_session() as session:
        project_id = _create_project_with_snapshot(session)
        session.add(
            Cabling(
                project_id=project_id,
                a_device="sw1",
                a_port="ge-0/0/1",
                a_port_type="dcim.interface",
                b_device="srv1",
                b_port="eth0",
                b_port_type="dcim.interface",
                cable_type="cat6",
                label="N-100",
                normalized_key="srv1::eth0::dcim.interface|sw1::ge-0/0/1::dcim.interface",
            )
        )
        session.commit()

        result = generate(session, project_id, "excel", None, "field-pack")
        session.commit()

        excel_file = (
            session.query(ArtifactFile)
            .filter(
                ArtifactFile.artifact_version_id == result.artifact_version.id,
                ArtifactFile.artifact_type == "excel",
            )
            .one()
        )
        wb = load_workbook(tmp_path / excel_file.relative_path)

        assert "work_steps" in wb.sheetnames
        assert "verification_checklist" in wb.sheetnames
        assert "issue_log" in wb.sheetnames

        work_steps = wb["work_steps"]
        assert work_steps.max_row >= 4
        assert work_steps.cell(row=1, column=1).value == "step_no"
        assert work_steps.cell(row=1, column=2).value == "depends_on_step_no"
        actions = [
            work_steps.cell(row=i, column=4).value
            for i in range(2, work_steps.max_row + 1)
        ]
        assert any(
            action and "Connect network cable N-100" in action for action in actions
        )
        assert any(action and "Location:" in action for action in actions)
        assert work_steps.cell(row=2, column=2).value in ("", None)
        assert work_steps.cell(row=3, column=2).value == 1

        verification = wb["verification_checklist"]
        assert verification.max_row >= 4
        assert verification.cell(row=1, column=3).value == "result"
        assert verification.cell(row=1, column=6).value == "step_no"

        issue_log = wb["issue_log"]
        assert issue_log.cell(row=1, column=4).value == "issue"


def test_template_categories_map_to_operation_phases_and_priority(
    tmp_path: Path,
) -> None:
    os.environ["RACKWRIGHT_DATA_DIR"] = str(tmp_path)
    with _new_session() as session:
        template_set = TemplateSet(name="phase-ts")
        session.add(template_set)
        session.flush()
        session.add_all(
            [
                TemplateSection(
                    template_set_id=template_set.id,
                    target_type="Project",
                    category="Cutover",
                    section_order=1,
                    output_targets='["excel"]',
                    applicable_roles=None,
                    text="execute cutover sequence",
                ),
                TemplateSection(
                    template_set_id=template_set.id,
                    target_type="Project",
                    category="Power",
                    section_order=1,
                    output_targets='["excel"]',
                    applicable_roles=None,
                    text="verify pdu feed",
                ),
                TemplateSection(
                    template_set_id=template_set.id,
                    target_type="Project",
                    category="Preconditions",
                    section_order=1,
                    output_targets='["excel"]',
                    applicable_roles=None,
                    text="confirm permit",
                ),
            ]
        )
        session.flush()

        project = create_project_from_template_set(
            session,
            project_name="phase-project",
            owner=None,
            notes=None,
            template_set_id=template_set.id,
        )
        session.commit()

        result = generate(session, project.id, "excel", None, "phase-run")
        session.commit()

        excel_file = (
            session.query(ArtifactFile)
            .filter(
                ArtifactFile.artifact_version_id == result.artifact_version.id,
                ArtifactFile.artifact_type == "excel",
            )
            .one()
        )
        wb = load_workbook(tmp_path / excel_file.relative_path)
        work_steps = wb["work_steps"]

        template_rows = [
            (
                work_steps.cell(row=i, column=3).value,
                work_steps.cell(row=i, column=4).value,
            )
            for i in range(2, work_steps.max_row + 1)
            if work_steps.cell(row=i, column=4).value
            and (
                "confirm permit" in work_steps.cell(row=i, column=4).value
                or "verify pdu feed" in work_steps.cell(row=i, column=4).value
                or "execute cutover sequence" in work_steps.cell(row=i, column=4).value
            )
        ]

        assert template_rows == [
            ("pre-check", "confirm permit"),
            ("power-execution", "verify pdu feed"),
            ("cutover", "execute cutover sequence"),
        ]
